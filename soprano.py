import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

caminho_arquivo1 = r'C:\asddsasdasd\Soprano1.xlsx'
caminho_arquivo2 = r'C:\asddsasdasd\Soprano2.xlsx'
caminho_salvamento = r'C:\asddsasdasd\resultado_soprano.xlsx'

arquivo1 = pd.read_excel(caminho_arquivo1, engine='openpyxl')
arquivo2 = pd.read_excel(caminho_arquivo2, engine='openpyxl')

arquivo1.dropna(how='all', inplace=True)
arquivo2.dropna(how='all', inplace=True)

colunas_comparacao = [
    "N Nota", "Codigo Produto", "NCM Produto", "CFOP Produto", 
    "Valor Produto", "Valor ICMS"
]

for coluna in colunas_comparacao:
    if coluna not in arquivo1.columns:
        arquivo1[coluna] = "NO DATA"

arquivo1.sort_values(by=["N Nota", "Codigo Produto"], inplace=True)
arquivo2.sort_values(by=["N Nota", "Codigo Produto"], inplace=True)

resultado_df = arquivo2.copy()

def comparar_linhas(row):
    
    Soprano1 = arquivo1[
        (arquivo1['N Nota'] == row['N Nota']) & 
        (arquivo1['Codigo Produto'] == row['Codigo Produto'])
    ]
    
    if Soprano1.empty:
        return "NO DATA", "Linha não encontrada em Soprano1", "Verificar origem dos dados"
    
    colunas_diferentes = []
    for coluna in colunas_comparacao:
        if row[coluna] != Soprano1.iloc[0][coluna]:
            colunas_diferentes.append(coluna)
    
    if not colunas_diferentes:
        return "TRUE", "Todas as colunas iguais", "Nenhuma ação necessária"
    
    return "FALSE", ", ".join(colunas_diferentes), "Verificar discrepâncias"

resultado_df[['Status Comparacao', 'Detalhes Diferenças', 'Ações Recomendadas']] = resultado_df.apply(
    lambda row: pd.Series(comparar_linhas(row)), axis=1
)

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(resultado_df, index=False, header=True):
    ws.append(r)

vermelho = Font(color="FF0000")
verde = Font(color="006400")
preto = Font(color="000000")
fundo_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fundo_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    status = row[-3].value
    detalhes = row[-2].value
    
    if status == "FALSE":
        row[-3].font = vermelho
        for cell in row:
            cell.fill = fundo_vermelho
    elif status == "TRUE":
        row[-3].font = verde
        for cell in row:
            cell.fill = fundo_verde
    elif status == "NO DATA":
        row[-3].font = preto
    
    if detalhes not in ["Todas as colunas iguais", "Linha não encontrada em Soprano1"]:
        colunas_diferentes = detalhes.split(", ")
        
        colunas_indices = [
            idx + 1 for idx, col in enumerate(resultado_df.columns) if col in colunas_diferentes
        ]
        
        for cell in row:
            coluna_letra = get_column_letter(cell.column)

            if cell.column in colunas_indices:
                cell.font = vermelho

                Soprano2_value = cell.value if cell.value is not None else 'N/A'

                n_nota = row[0].value
                codigo_produto = row[1].value
                
                if n_nota is not None and codigo_produto is not None:
                    n_nota = str(n_nota).strip()
                    codigo_produto = str(codigo_produto).strip()
                    
                    Soprano1_row = arquivo1[
                        (arquivo1['N Nota'].astype(str).str.strip() == n_nota) &
                        (arquivo1['Codigo Produto'].astype(str).str.strip() == codigo_produto)
                    ]

                    if not Soprano1_row.empty:
                        Soprano1_value = Soprano1_row.iloc[0][resultado_df.columns[cell.column - 1]]
                    else:
                        Soprano1_value = 'N/A'
                else:
                    Soprano1_value = 'N/A'

                comentario = f"Soprano1: {Soprano1_value} | Soprano2: {Soprano2_value}"
                cell.comment = Comment(comentario, "Autor")


ws.auto_filter.ref = ws.dimensions

header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
header_font = Font(bold=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

ws_resumo = wb.create_sheet(title="Resumo")

contagem_status = resultado_df['Status Comparacao'].value_counts()
ws_resumo.append(["Status", "Quantidade"])
for status, quantidade in contagem_status.items():
    ws_resumo.append([status, quantidade])

grafico = BarChart()
dados = Reference(ws_resumo, min_col=2, min_row=1, max_row=len(contagem_status) + 1, max_col=2)
categorias = Reference(ws_resumo, min_col=1, min_row=2, max_row=len(contagem_status) + 1)
grafico.add_data(dados, titles_from_data=True)
grafico.set_categories(categorias)
grafico.title = "Distribuição de Status"
grafico.y_axis.title = "Quantidade"
grafico.x_axis.title = "Status"

ws_resumo.add_chart(grafico, "D2")

wb.save(caminho_salvamento)
print(f"Arquivo salvo com sucesso em: {caminho_salvamento}")