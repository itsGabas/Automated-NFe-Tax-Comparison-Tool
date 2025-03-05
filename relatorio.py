import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, Reference

# Caminho dos arquivos
caminho_arquivo1 = r'C:\asddsasdasd\JS1.xlsx'  # Arquivo verdadeiro
caminho_arquivo2 = r'C:\asddsasdasd\JS2.xlsx'  # Arquivo a ser comparado
caminho_salvamento = r'C:\asddsasdasd\resultado_comparacao.xlsx'  # Arquivo final

# Ler os arquivos
arquivo1 = pd.read_excel(caminho_arquivo1, engine='openpyxl')
arquivo2 = pd.read_excel(caminho_arquivo2, engine='openpyxl')

# Garantir que apenas linhas completamente vazias sejam removidas
arquivo1.dropna(how='all', inplace=True)
arquivo2.dropna(how='all', inplace=True)

# Colunas para comparar
colunas_comparacao = [
    "N Nota", "Codigo Produto", "NCM Produto", "CFOP Produto", 
    "Valor Produto", "CST ICMS", "Valor ICMS"
]

# Garantir que todas as colunas existam no JS1, preenchendo com 'NO DATA' onde necessário
for coluna in colunas_comparacao:
    if coluna not in arquivo1.columns:
        arquivo1[coluna] = "NO DATA"

# Ordenar os DataFrames pelas colunas de referência
arquivo1.sort_values(by=["N Nota", "Codigo Produto"], inplace=True)
arquivo2.sort_values(by=["N Nota", "Codigo Produto"], inplace=True)

# Criar DataFrame final com todas as colunas do JS2
resultado_df = arquivo2.copy()

# Função para comparar as linhas
def comparar_linhas(row):
    # Encontrar a linha correspondente em JS1
    js1_row = arquivo1[
        (arquivo1['N Nota'] == row['N Nota']) & 
        (arquivo1['Codigo Produto'] == row['Codigo Produto'])
    ]
    
    # Se não houver correspondência, retornar "NO DATA" para status e detalhes
    if js1_row.empty:
        return "NO DATA", "Linha não encontrada em JS1", "Verificar origem dos dados"
    
    # Verificar quais colunas estão diferentes
    colunas_diferentes = []
    for coluna in colunas_comparacao:
        if row[coluna] != js1_row.iloc[0][coluna]:
            colunas_diferentes.append(coluna)
    
    # Se não houver diferenças, retornar "TRUE" e uma mensagem vazia
    if not colunas_diferentes:
        return "TRUE", "Todas as colunas iguais", "Nenhuma ação necessária"
    
    # Se houver diferenças, retornar "FALSE" e as colunas diferentes
    return "FALSE", ", ".join(colunas_diferentes), "Verificar discrepâncias"

# Aplicar a função de comparação a cada linha
resultado_df[['Status Comparacao', 'Detalhes Diferenças', 'Ações Recomendadas']] = resultado_df.apply(
    lambda row: pd.Series(comparar_linhas(row)), axis=1
)

# Salvar o DataFrame em um arquivo Excel
resultado_df.to_excel(caminho_salvamento, index=False)

# Aplicar formatação condicional usando openpyxl
wb = Workbook()
ws = wb.active

# Copiar o DataFrame para a planilha do openpyxl
for r in dataframe_to_rows(resultado_df, index=False, header=True):
    ws.append(r)

# Definir as cores
vermelho = Font(color="FF0000")  # Vermelho para FALSE e diferenças
verde = Font(color="006400")     # Verde escuro para TRUE
preto = Font(color="000000")     # Preto para NO DATA
fundo_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fundo_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro para TRUE

# Aplicar formatação condicional
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    status = row[-3].value  # Status Comparacao está na antepenúltima coluna
    detalhes = row[-2].value  # Detalhes Diferenças está na penúltima coluna
    
    # Formatar o status e o fundo da linha
    if status == "FALSE":
        row[-3].font = vermelho
        for cell in row:
            cell.fill = fundo_vermelho  # Destacar a linha inteira em vermelho claro
    elif status == "TRUE":
        row[-3].font = verde
        for cell in row:
            cell.fill = fundo_verde  # Destacar a linha inteira em verde claro
    elif status == "NO DATA":
        row[-3].font = preto
    
    # Formatar as células com diferenças
    if detalhes != "Todas as colunas iguais" and detalhes != "Linha não encontrada em JS1":
        colunas_diferentes = detalhes.split(", ")
        for cell in row:
            if cell.column_letter in [ws.cell(row=1, column=idx + 1).column_letter for idx, col in enumerate(resultado_df.columns) if col in colunas_diferentes]:
                cell.font = vermelho

# Adicionar filtros automáticos
ws.auto_filter.ref = ws.dimensions

# Formatar cabeçalho
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
header_font = Font(bold=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

# Criar uma nova planilha para o gráfico
ws_resumo = wb.create_sheet(title="Resumo")

# Contar a quantidade de cada status
contagem_status = resultado_df['Status Comparacao'].value_counts()
ws_resumo.append(["Status", "Quantidade"])
for status, quantidade in contagem_status.items():
    ws_resumo.append([status, quantidade])

# Criar o gráfico de barras
grafico = BarChart()
dados = Reference(ws_resumo, min_col=2, min_row=1, max_row=len(contagem_status) + 1, max_col=2)
categorias = Reference(ws_resumo, min_col=1, min_row=2, max_row=len(contagem_status) + 1)
grafico.add_data(dados, titles_from_data=True)
grafico.set_categories(categorias)
grafico.title = "Distribuição de Status"
grafico.y_axis.title = "Quantidade"
grafico.x_axis.title = "Status"

# Adicionar o gráfico à planilha de resumo
ws_resumo.add_chart(grafico, "D2")

# Salvar o arquivo com a formatação
wb.save(caminho_salvamento)
print(f"Arquivo salvo com sucesso em: {caminho_salvamento}")